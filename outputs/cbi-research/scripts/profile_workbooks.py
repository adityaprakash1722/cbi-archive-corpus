#!/usr/bin/env python3
"""Inventory every CBI XLS/XLSX workbook without mutating source files."""

from __future__ import annotations

import argparse
import csv
import io
import json
import re
import sys
import time
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


def arguments() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--archive", type=Path, required=True)
    parser.add_argument("--catalog", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--xlrd-path", type=Path)
    parser.add_argument("--progress-every", type=int, default=25)
    return parser.parse_args()


def read_catalog(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as stream:
        return [row for row in csv.DictReader(stream) if row["format"] in {"XLS", "XLSX"}]


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as stream:
        writer = csv.DictWriter(stream, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def classify(url: str) -> tuple[str, str]:
    value = url.lower()
    filename = value.rsplit("/", 1)[-1].split("?", 1)[0]
    if any(
        token in value
        for token in (
            "chartpack",
            "/publications/quarterly-bulletins/",
            "/research-exchange/staff-insights/",
            "/financial-conditions-of-credit-unions/",
        )
    ):
        return "research_supporting_data", "Research/publication supporting data or chartpack"
    if any(token in value for token in ("reporting-template", "reporting_template", "return-template")):
        return "regulatory_reporting_template", "URL identifies a reporting/return template"
    if "/reporting-requirements/" in value or "/statistical-reporting-requirements/" in value:
        return "regulatory_reporting_template", "Located under reporting requirements"
    if any(token in filename for token in ("application", "questionnaire", "sign-off", "submission-form")):
        return "application_or_form", "Filename identifies an application/form/questionnaire"
    if "/forms" in value or "-form." in filename or "_form." in filename:
        return "application_or_form", "Located in forms or filename identifies a form"
    if any(token in value for token in ("beneficial-ownership-form", "self-reporting", "insider-list")):
        return "application_or_form", "URL identifies a form or submission list"
    if any(
        token in value
        for token in (
            "taxonomy", "validation", "business-rules", "known-issues", "dpm_dictionary",
            "dpm-dictionary", "schema_excel", "data-quality-checks",
        )
    ):
        return "regulatory_taxonomy_or_validation", "URL identifies taxonomy, validation or business-rule metadata"
    if any(token in value for token in ("outsourcing-register", "/resolution/bifr/", "daofi-mcr")):
        return "regulatory_reporting_template", "URL identifies a regulatory register or return"
    if any(
        token in value
        for token in (
            "covered-bond-programmes", "asset-covered-securities", "net-short-positions",
            "/regulation/psd2/article-",
        )
    ):
        return "official_list_or_disclosure", "URL identifies a published regulatory list/disclosure"
    if any(
        token in value
        for token in (
            "/statistics/data-and-analysis/",
            "/statistics/interest-rates-exchange-rates/",
            "/statistics/statistical-publications/",
            "/statistics/frontier-statistics/",
            "/financial-system/access-to-cash/",
            "public-data",
            "publishing-dashboard",
            "levels-of-compliance",
            "securities-lending-isins",
        )
    ):
        return "published_data", "Located in a published-data area"
    if any(token in filename for token in ("template", "return", "schedule")):
        return "regulatory_reporting_template", "Filename identifies a template/return/schedule"
    return "other_workbook", "No deterministic URL classification matched"


def sample_openpyxl_sheet(sheet: Any, row_limit: int = 25, col_limit: int = 15) -> dict[str, Any]:
    nonempty = 0
    numeric = 0
    formulas = 0
    texts: list[str] = []
    for row in sheet.iter_rows(
        min_row=1,
        max_row=min(sheet.max_row or 1, row_limit),
        min_col=1,
        max_col=min(sheet.max_column or 1, col_limit),
        values_only=False,
    ):
        for cell in row:
            value = cell.value
            if value in {None, ""}:
                continue
            nonempty += 1
            if isinstance(value, (int, float)):
                numeric += 1
            if isinstance(value, str) and value.startswith("="):
                formulas += 1
            if isinstance(value, str) and not value.startswith("=") and len(texts) < 12:
                cleaned = re.sub(r"\s+", " ", value).strip()
                if cleaned:
                    texts.append(cleaned[:160])
    return {"sample_nonempty": nonempty, "sample_numeric": numeric, "sample_formulas": formulas, "sample_text": " | ".join(texts)}


def inspect_xlsx(path: Any) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(
        path, read_only=True, data_only=False, keep_links=False
    )
    sheets = []
    try:
        for sheet in workbook.worksheets:
            sample = sample_openpyxl_sheet(sheet)
            sheets.append({
                "name": sheet.title,
                "state": sheet.sheet_state,
                "rows": sheet.max_row or 0,
                "columns": sheet.max_column or 0,
                "apparent_cells": (sheet.max_row or 0) * (sheet.max_column or 0),
                **sample,
            })
    finally:
        workbook.close()
    return {"sheets": sheets}


def inspect_xls(path: Any, xlrd: Any, *, file_contents: bytes | None = None) -> dict[str, Any]:
    if file_contents is None:
        workbook = xlrd.open_workbook(path, on_demand=True, formatting_info=False)
    else:
        workbook = xlrd.open_workbook(
            file_contents=file_contents, on_demand=True, formatting_info=False
        )
    sheets = []
    try:
        for name in workbook.sheet_names():
            sheet = workbook.sheet_by_name(name)
            nonempty = 0
            numeric = 0
            texts: list[str] = []
            for row_index in range(min(sheet.nrows, 25)):
                for col_index in range(min(sheet.ncols, 15)):
                    value = sheet.cell_value(row_index, col_index)
                    if value in {None, ""}:
                        continue
                    nonempty += 1
                    if isinstance(value, (int, float)):
                        numeric += 1
                    elif len(texts) < 12:
                        cleaned = re.sub(r"\s+", " ", str(value)).strip()
                        if cleaned:
                            texts.append(cleaned[:160])
            sheets.append({
                "name": name,
                "state": "unknown",
                "rows": sheet.nrows,
                "columns": sheet.ncols,
                "apparent_cells": sheet.nrows * sheet.ncols,
                "sample_nonempty": nonempty,
                "sample_numeric": numeric,
                "sample_formulas": None,
                "sample_text": " | ".join(texts),
            })
            workbook.unload_sheet(name)
    finally:
        workbook.release_resources()
    return {"sheets": sheets}


def inspect_encrypted(path: Path, xlrd: Any) -> dict[str, Any]:
    import msoffcrypto  # type: ignore

    failures = []
    for password in ("VelvetSweatshop", ""):
        try:
            with path.open("rb") as source:
                office = msoffcrypto.OfficeFile(source)
                office.load_key(password=password)
                decrypted = io.BytesIO()
                office.decrypt(decrypted)
            payload = decrypted.getvalue()
            if payload.startswith(b"PK\x03\x04"):
                details = inspect_xlsx(io.BytesIO(payload))
                details["detected_format"] = "XLSX"
            else:
                details = inspect_xls(path, xlrd, file_contents=payload)
                details["detected_format"] = "XLS"
            details["decryption"] = f"known default password: {password or '<blank>'}"
            return details
        except Exception as exc:
            failures.append(f"{password or '<blank>'}: {type(exc).__name__}: {exc}")
    raise RuntimeError("Known-password decryption failed; " + " | ".join(failures))


def main() -> int:
    args = arguments()
    if args.xlrd_path:
        sys.path.insert(0, str(args.xlrd_path.resolve()))
    import xlrd  # type: ignore

    archive = args.archive.resolve()
    output = args.output.resolve()
    catalog = read_catalog(args.catalog.resolve())
    workbook_rows: list[dict[str, Any]] = []
    sheet_rows: list[dict[str, Any]] = []
    for index, entry in enumerate(catalog, 1):
        path = archive / entry["local_path"].replace("\\", "/")  # manifest paths are written on Windows
        category, reason = classify(entry["canonical_url"])
        started = time.monotonic()
        status = "readable"
        error = ""
        details: dict[str, Any] = {"sheets": [], "detected_format": entry["format"], "decryption": ""}
        try:
            if entry["format"] == "XLSX":
                details = inspect_xlsx(path)
                details["detected_format"] = "XLSX"
                details["decryption"] = ""
            else:
                with path.open("rb") as prefix_stream:
                    prefix = prefix_stream.read(4)
                if prefix == b"PK\x03\x04":
                    details = inspect_xlsx(io.BytesIO(path.read_bytes()))
                    details["detected_format"] = "XLSX"
                    details["decryption"] = ""
                else:
                    try:
                        details = inspect_xls(path, xlrd)
                        details["detected_format"] = "XLS"
                        details["decryption"] = ""
                    except Exception as exc:
                        if "encrypted" not in str(exc).lower():
                            raise
                        details = inspect_encrypted(path, xlrd)
        except Exception as exc:  # preserve the failed file in the inventory
            status = "unreadable"
            error = f"{type(exc).__name__}: {exc}"[:1000]

        sheets = details["sheets"]
        total_cells = sum(sheet["apparent_cells"] for sheet in sheets)
        sampled_nonempty = sum(sheet["sample_nonempty"] for sheet in sheets)
        sampled_numeric = sum(sheet["sample_numeric"] for sheet in sheets)
        sampled_formulas = sum(sheet["sample_formulas"] or 0 for sheet in sheets)
        workbook_rows.append({
            "sha256": entry["sha256"],
            "format": entry["format"],
            "detected_format": details.get("detected_format", ""),
            "decryption": details.get("decryption", ""),
            "bytes": int(entry["bytes"]),
            "classification": category,
            "classification_reason": reason,
            "read_status": status,
            "sheet_count": len(sheets),
            "visible_sheet_count": sum(sheet["state"] in {"visible", "unknown"} for sheet in sheets),
            "apparent_cells": total_cells,
            "sample_nonempty": sampled_nonempty,
            "sample_numeric": sampled_numeric,
            "sample_formulas": sampled_formulas,
            "sheet_names": " | ".join(sheet["name"] for sheet in sheets),
            "sample_text": " || ".join(sheet["sample_text"] for sheet in sheets if sheet["sample_text"])[:4000],
            "seconds": round(time.monotonic() - started, 3),
            "error": error,
            "canonical_url": entry["canonical_url"],
            "local_path": entry["local_path"],
        })
        for sheet in sheets:
            sheet_rows.append({
                "sha256": entry["sha256"],
                "format": entry["format"],
                "classification": category,
                "sheet_name": sheet["name"],
                "state": sheet["state"],
                "rows": sheet["rows"],
                "columns": sheet["columns"],
                "apparent_cells": sheet["apparent_cells"],
                "sample_nonempty": sheet["sample_nonempty"],
                "sample_numeric": sheet["sample_numeric"],
                "sample_formulas": sheet["sample_formulas"],
                "sample_text": sheet["sample_text"],
                "canonical_url": entry["canonical_url"],
            })
        if index % args.progress_every == 0 or index == len(catalog):
            print(f"Inspected {index}/{len(catalog)} workbooks", flush=True)

    fields = [
        "sha256", "format", "detected_format", "decryption", "bytes", "classification", "classification_reason", "read_status",
        "sheet_count", "visible_sheet_count", "apparent_cells", "sample_nonempty",
        "sample_numeric", "sample_formulas", "sheet_names", "sample_text", "seconds", "error",
        "canonical_url", "local_path",
    ]
    write_csv(output / "workbook-profile.csv", workbook_rows, fields)
    write_csv(
        output / "workbook-sheet-profile.csv",
        sheet_rows,
        [
            "sha256", "format", "classification", "sheet_name", "state", "rows", "columns",
            "apparent_cells", "sample_nonempty", "sample_numeric", "sample_formulas", "sample_text",
            "canonical_url",
        ],
    )
    summary = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "logical_workbooks": len(workbook_rows),
        "by_format": dict(sorted(Counter(row["format"] for row in workbook_rows).items())),
        "by_classification": dict(sorted(Counter(row["classification"] for row in workbook_rows).items())),
        "read_status": dict(sorted(Counter(row["read_status"] for row in workbook_rows).items())),
        "total_sheets": len(sheet_rows),
        "hidden_sheets": sum(row["state"] not in {"visible", "unknown"} for row in sheet_rows),
        "caveat": (
            "Classification is deterministic from URL/filename. Dimensions are workbook metadata; "
            "apparent cell counts may include formatted but empty ranges. Samples cover at most the "
            "first 25 rows by 15 columns of each sheet."
        ),
    }
    (output / "workbook-profile-summary.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    print(json.dumps(summary, indent=2, ensure_ascii=False), flush=True)
    return 0 if summary["read_status"].get("unreadable", 0) == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
